from openpyxl import Workbook, load_workbook
import os

def write_to_excel(data, file_path):
    if not isinstance(file_path, str):
        print(f"Invalid file path in write_to_excel: {file_path} (type: {type(file_path)})")
        return

    print(f"Attempting to write data to {file_path}")

    try:
        if os.path.exists(file_path):
            print("File exists. Loading existing workbook.")
            book = load_workbook(file_path)
            sheet = book.active
        else:
            print("File does not exist. Creating new workbook.")
            book = Workbook()
            sheet = book.active
            sheet.title = 'Sheet1'

        if not sheet.max_row > 1:
            headers = ['name', 'phone_no', 'email_id', 'designation', 'department']
            sheet.append(headers)

        for entry in data:
            row = [
                str(entry.get('name', '')),
                str(entry.get('phone_no', '')),
                str(entry.get('email_id', '')),
                str(entry.get('designation', '')),
                str(entry.get('department', ''))
            ]
            sheet.append(row)

        book.save(file_path)
        print("Data successfully written to Excel.")
    except Exception as e:
        print(f"Failed to write data to Excel: {e}")

def read_from_excel(file_path):
    if not isinstance(file_path, str):
        print(f"Invalid file path in read_from_excel: {file_path} (type: {type(file_path)})")
        return []

    print(f"Attempting to read data from {file_path}")

    if not os.path.exists(file_path):
        print(f"File does not exist in read_from_excel: {file_path}")
        return []

    try:
        book = load_workbook(file_path)
        sheet = book.active
        data = []

        print(f"Sheet max_row: {sheet.max_row}, max_column: {sheet.max_column}")

        for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if len(row) < 5:
                print(f"Skipping incomplete row: {row}")
                continue

            data.append({
                'id': index,  # Using index as a temporary unique identifier if ID is missing
                'name': row[0],
                'phone_no': row[1],
                'email_id': row[2],
                'designation': row[3],
                'department': row[4]
            })

        return data
    except Exception as e:
        print(f"Failed to read data from Excel: {e}")
        return []

def delete_from_excel(file_path, row_id):
    if not isinstance(file_path, str):
        print(f"Invalid file path in delete_from_excel: {file_path} (type: {type(file_path)})")
        return False

    if not isinstance(row_id, int):
        print(f"Invalid row_id in delete_from_excel: {row_id} (type: {type(row_id)})")
        return False

    if not os.path.exists(file_path):
        print(f"File does not exist in delete_from_excel: {file_path}")
        return False

    try:
        book = load_workbook(file_path)
        sheet = book.active

        # Adjust for zero-based index
        if row_id < 2 or row_id > sheet.max_row:
            print(f"Row with ID {row_id} not found in {file_path}.")
            return False

        sheet.delete_rows(row_id)
        book.save(file_path)
        print(f"Successfully deleted row with ID {row_id} from {file_path}.")
        return True
    except Exception as e:
        print(f"Failed to delete data from Excel: {e}")
        return False

def update_excel(file_path, row_id, updated_data):
    if not isinstance(file_path, str):
        print(f"Invalid file path in update_excel: {file_path} (type: {type(file_path)})")
        return False
    
    if not isinstance(row_id, int):
        print(f"Invalid row_id in update_excel: {row_id} (type: {type(row_id)})")
        return False

    if not os.path.exists(file_path):
        print(f"File does not exist in update_excel: {file_path}")
        return False

    try:
        book = load_workbook(file_path)
        sheet = book.active

        # Adjust for one-based index (Excel row numbering starts from 1)
        if row_id < 2 or row_id > sheet.max_row:
            print(f"Row with ID {row_id} not found in {file_path}.")
            return False

        for key, value in updated_data.items():
            if key in ['name', 'phone_no', 'email_id', 'department']:
                col_idx = ['name', 'phone_no', 'email_id', 'department'].index(key) + 1
                sheet.cell(row=row_id, column=col_idx).value = value

        book.save(file_path)
        print(f"Successfully updated row with ID {row_id} in {file_path}.")
        return True
    except Exception as e:
        print(f"Failed to update data in Excel: {e}")
        return False
