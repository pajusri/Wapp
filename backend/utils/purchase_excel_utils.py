from openpyxl import Workbook, load_workbook
import os

def write_to_excel(data, file_path, sheet_name):
    if not isinstance(file_path, str):
        print(f"Invalid file path in write_to_excel: {file_path} (type: {type(file_path)})")
        return

    print(f"Attempting to write data to {file_path} in sheet {sheet_name}")

    try:
        if os.path.exists(file_path):
            print("File exists. Loading existing workbook.")
            book = load_workbook(file_path)
            if sheet_name in book.sheetnames:
                sheet = book[sheet_name]
            else:
                sheet = book.create_sheet(title=sheet_name)
        else:
            print("File does not exist. Creating new workbook.")
            book = Workbook()
            sheet = book.active
            sheet.title = sheet_name

        # Add headers if the sheet is new
        if sheet.max_row == 1 and sheet.max_column == 1:
            headers = ['market', 'lot_no', 'kgs_purchased', 'price_per_kg', 'farmer_name', 'date_of_purchase', 'total_cost']
            sheet.append(headers)

        for entry in data:
            row = [
                entry.get('market', ''),
                entry.get('lot_no', ''),
                entry.get('kgs_purchased', ''),
                entry.get('price_per_kg', ''),
                entry.get('farmer_name', ''),
                entry.get('date_of_purchase', ''),
                entry.get('total_cost', '')
            ]
            sheet.append(row)

        book.save(file_path)
        print("Data successfully written to Excel.")
    except Exception as e:
        print(f"Failed to write data to Excel: {e}")

def read_from_excel(file_path, sheet_name):
    if not isinstance(file_path, str):
        print(f"Invalid file path in read_from_excel: {file_path} (type: {type(file_path)})")
        return []

    print(f"Attempting to read data from {file_path} in sheet {sheet_name}")

    if not os.path.exists(file_path):
        print(f"File does not exist in read_from_excel: {file_path}")
        return []

    try:
        book = load_workbook(file_path)
        if sheet_name not in book.sheetnames:
            print(f"Sheet {sheet_name} does not exist in {file_path}")
            return []

        sheet = book[sheet_name]
        data = []

        for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if len(row) < 7:
                continue
            data.append({
                'id': index,
                'market': row[0],
                'lot_no': row[1],
                'kgs_purchased': row[2],
                'price_per_kg': row[3],
                'farmer_name': row[4],
                'date_of_purchase': row[5],
                'total_cost': row[6]
            })

        return data
    except Exception as e:
        print(f"Failed to read data from Excel: {e}")
        return []

def delete_purchase_from_excel(file_path, sheet_name, row_id):
    if not isinstance(file_path, str):
        print(f"Invalid file path in delete_purchase_from_excel: {file_path} (type: {type(file_path)})")
        return False

    if not isinstance(row_id, int):
        print(f"Invalid row_id in delete_purchase_from_excel: {row_id} (type: {type(row_id)})")
        return False

    if not os.path.exists(file_path):
        print(f"File does not exist in delete_purchase_from_excel: {file_path}")
        return False

    try:
        book = load_workbook(file_path)
        if sheet_name not in book.sheetnames:
            print(f"Sheet {sheet_name} does not exist in {file_path}")
            return False

        sheet = book[sheet_name]
        if row_id < 2 or row_id > sheet.max_row:
            print(f"Row with ID {row_id} not found in sheet {sheet_name}.")
            return False

        sheet.delete_rows(row_id)
        book.save(file_path)
        print(f"Successfully deleted row with ID {row_id} from sheet {sheet_name}.")
        return True
    except Exception as e:
        print(f"Failed to delete data from Excel: {e}")
        return False
