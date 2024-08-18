from openpyxl import Workbook, load_workbook
import os

def write_to_excel(data, file_path):
    if not isinstance(file_path, str):
        print(f"Invalid file path in write_to_excel: {file_path} (type: {type(file_path)})")
        return

    print(f"Attempting to write data to {file_path}")

    try:
        if os.path.exists(file_path):
            print("File exists. Loading existing workbook.")
            book = load_workbook(file_path)
            sheet = book.active
        else:
            print("File does not exist. Creating new workbook.")
            book = Workbook()
            sheet = book.active
            sheet.title = 'Sheet1'
            sheet.append(['id', 'name', 'phone_no', 'email_id', 'chakki_center_name', 'address'])  # Include 'id' as a header

        # Determine the next available ID
        if sheet.max_row > 1:
            last_id = sheet.cell(row=sheet.max_row, column=1).value
        else:
            last_id = 0

        for entry in data:
            last_id += 1
            row = [last_id, entry.get('name', ''), entry.get('phone_no', ''), entry.get('email_id', ''), entry.get('chakki_center_name', ''), entry.get('address', '')]
            sheet.append(row)

        book.save(file_path)
        print(f"Data successfully written to Excel with next ID {last_id}.")
    except Exception as e:
        print(f"Failed to write data to Excel: {e}")

def read_from_excel(file_path):
    if not isinstance(file_path, str):
        print(f"Invalid file path in read_from_excel: {file_path} (type: {type(file_path)})")
        return []

    print(f"Attempting to read data from {file_path}")

    if not os.path.exists(file_path):
        print(f"File does not exist in read_from_excel: {file_path}")
        return []

    try:
        book = load_workbook(file_path)
        sheet = book.active
        data = []

        print(f"Sheet max_row: {sheet.max_row}, max_column: {sheet.max_column}")

        for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if len(row) < 5:
                print(f"Skipping incomplete row: {row}")
                continue

            data.append({
                'id': index,  # Using index as a temporary unique identifier if ID is missing
                'name': row[0],
                'phone_no': row[1],
                'email_id': row[2],
                'chakki_center_name': row[3],
                'address': row[4]
            })

        #print(f"Retrieved customer data: {data}")
        return data
    except Exception as e:
        print(f"Failed to read data from Excel: {e}")
        return []

def delete_from_excel(file_path, row_id):
    if not isinstance(file_path, str):
        print(f"Invalid file path in delete_from_excel: {file_path} (type: {type(file_path)})")
        return False

    if not isinstance(row_id, int):
        print(f"Invalid row_id in delete_from_excel: {row_id} (type: {type(row_id)})")
        return False

    if not os.path.exists(file_path):
        print(f"File does not exist in delete_from_excel: {file_path}")
        return False

    try:
        book = load_workbook(file_path)
        sheet = book.active

        # Adjust for zero-based index
        if row_id < 2 or row_id > sheet.max_row:
            print(f"Row with ID {row_id} not found in {file_path}.")
            return False

        sheet.delete_rows(row_id)
        book.save(file_path)
        print(f"Successfully deleted row with ID {row_id} from {file_path}.")
        return True
    except Exception as e:
        print(f"Failed to delete data from Excel: {e}")
        return False

def update_excel(file_path, row_id, updated_data):
    if not isinstance(file_path, str):
        print(f"Invalid file path in update_excel: {file_path} (type: {type(file_path)})")
        return False

    if not isinstance(row_id, int):
        print(f"Invalid row_id in update_excel: {row_id} (type: {type(row_id)})")
        return False

    if not os.path.exists(file_path):
        print(f"File does not exist in update_excel: {file_path}")
        return False

    try:
        book = load_workbook(file_path)
        sheet = book.active

        # Adjust for zero-based index
        if row_id < 2 or row_id > sheet.max_row:
            print(f"Row with ID {row_id} not found in {file_path}.")
            return False

        for key, value in updated_data.items():
            if key in ['name', 'phone_no', 'email_id', 'chakki_center_name', 'address']:
                col_idx = ['name', 'phone_no', 'email_id', 'chakki_center_name', 'address'].index(key) + 1
                sheet.cell(row=row_id, column=col_idx).value = value

        book.save(file_path)
        print(f"Successfully updated row with ID {row_id} in {file_path}.")
        return True
    except Exception as e:
        print(f"Failed to update data in Excel: {e}")
        return False
