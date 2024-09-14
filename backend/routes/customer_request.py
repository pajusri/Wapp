from flask import Blueprint, request, jsonify
import openpyxl
import os

customer_request_bp = Blueprint('customer_request', __name__)

# Path to Excel file
FILE_PATH = '/Users/admin/APP1/customer_requests.xlsx'

# Load or create Excel file and sheet
def load_or_create_workbook():
    try:
        if os.path.exists(FILE_PATH):
            workbook = openpyxl.load_workbook(FILE_PATH)
        else:
            workbook = openpyxl.Workbook()
            requests_sheet = workbook.active
            requests_sheet.title = "Requests"
            requests_sheet.append(['Chakki Center', 'Eggs Requested', 'Delivery Date', 'Hatching Date', 'Sent to Billing'])
            billing_sheet = workbook.create_sheet("Billing")
            billing_sheet.append(['Chakki Center', 'Eggs Requested', 'Delivery Date', 'Hatching Date'])
            workbook.save(FILE_PATH)
        return workbook
    except Exception as e:
        print(f"Error loading or creating workbook: {str(e)}")
        raise

# Create customer request
@customer_request_bp.route('/customer/request', methods=['POST'])
def customer_request():
    data = request.get_json()
    chakki_center = data.get('chakkiCenter')
    eggs_requested = data.get('eggsRequested')
    delivery_date = data.get('deliveryDate')
    hatching_date = data.get('hatchingDate')

    if not chakki_center or not eggs_requested or not delivery_date or not hatching_date:
        return jsonify({"error": "All fields are required"}), 400

    # Validate that delivery date is before hatching date
    if delivery_date >= hatching_date:
        return jsonify({"error": "Delivery date must be earlier than hatching date"}), 400

    # Load Excel and append the data
    try:
        workbook = load_or_create_workbook()
        sheet = workbook['Requests']
        sheet.append([chakki_center, eggs_requested, delivery_date, hatching_date, 'No'])
        workbook.save(FILE_PATH)
        return jsonify({"message": "Request successfully saved"}), 200
    except Exception as e:
        print(f"Error saving request: {str(e)}")  # Log the error
        return jsonify({"error": str(e)}), 500

# Get active requests
@customer_request_bp.route('/customer/requests', methods=['GET'])
def get_active_requests():
    try:
        workbook = load_or_create_workbook()
        sheet = workbook['Requests']
        active_requests = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[4] == 'No':  # Not sent to billing
                active_requests.append({
                    'chakkiCenter': row[0],
                    'eggsRequested': row[1],
                    'deliveryDate': row[2],
                    'hatchingDate': row[3]
                })

        return jsonify(active_requests), 200
    except Exception as e:
        print(f"Error fetching active requests: {str(e)}")  # Log the error
        return jsonify({'error': str(e)}), 500

# Send request to billing
@customer_request_bp.route('/customer/request/send-to-billing', methods=['POST'])
def send_to_billing():
    data = request.get_json()  # Make sure the request is sending JSON data
    chakki_center = data.get('chakkiCenter')

    try:
        workbook = load_or_create_workbook()
        requests_sheet = workbook['Requests']
        billing_sheet = workbook['Billing']

        # Find the entry in the Requests sheet and mark it as sent to billing
        found = False
        for row in requests_sheet.iter_rows(min_row=2):
            if row[0].value == chakki_center and row[4].value == 'No':  # Only move entries that are not sent to billing yet
                # Append the data to the Billing sheet
                billing_sheet.append([row[0].value, row[1].value, row[2].value, row[3].value])

                # Mark the request as sent to billing in the Requests sheet
                row[4].value = 'Yes'
                found = True
                print(f"Marked {chakki_center} as sent to billing and moved to the Billing sheet.")

        if found:
            workbook.save(FILE_PATH)
            return jsonify({"message": f"{chakki_center} sent to billing"}), 200
        else:
            return jsonify({"error": f"Entry for {chakki_center} not found or already sent to billing"}), 404
    except Exception as e:
        print(f"Error marking {chakki_center} as sent to billing: {str(e)}")
        return jsonify({'error': str(e)}), 500

# Get billing requests
@customer_request_bp.route('/customer/billing', methods=['GET'])
def get_billing_requests():
    try:
        workbook = load_or_create_workbook()
        billing_sheet = workbook['Billing']
        billing_requests = []

        for row in billing_sheet.iter_rows(min_row=2, values_only=True):
            billing_requests.append({
                'chakkiCenter': row[0],
                'eggsRequested': row[1],
                'deliveryDate': row[2],
                'hatchingDate': row[3]
            })

        return jsonify(billing_requests), 200
    except Exception as e:
        print(f"Error fetching billing requests: {str(e)}")  # Log the error
        return jsonify({'error': str(e)}), 500

# Delete customer request
@customer_request_bp.route('/customer/request/delete', methods=['POST'])
def delete_request():
    data = request.get_json()
    chakki_center = data.get('chakkiCenter')

    try:
        workbook = load_or_create_workbook()
        sheet = workbook['Requests']
        row_to_delete = None

        for row in sheet.iter_rows(min_row=2):
            if row[0].value == chakki_center:
                row_to_delete = row
                break

        if row_to_delete:
            sheet.delete_rows(row_to_delete[0].row)
            workbook.save(FILE_PATH)
            return jsonify({"message": f"Request for {chakki_center} deleted successfully."}), 200
        else:
            return jsonify({"error": f"Request for {chakki_center} not found."}), 404

    except Exception as e:
        print(f"Error deleting request: {str(e)}")
        return jsonify({'error': str(e)}), 500
