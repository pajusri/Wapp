from flask import Blueprint, request, jsonify
import openpyxl

# Define the blueprint
waste_cocoons_bp = Blueprint('waste_cocoons', __name__)

# File path for the Excel file
FILE_PATH = '/Users/admin/APP1/waste_cocoons_data.xlsx'


@waste_cocoons_bp.route('/waste_cocoons/entry', methods=['POST'])
def waste_cocoons_entry():
    data = request.get_json()
    breed = data.get('breed')
    kgs = data.get('kgs')

    if not breed or not kgs:
        return jsonify({"error": "Breed and kgs are required"}), 400

    try:
        # Load or create the Excel file
        try:
            workbook = openpyxl.load_workbook(FILE_PATH)
            sheet = workbook.active
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(['Breed', 'Kgs'])  # Add headers if the file doesn't exist

        # Add the new entry
        sheet.append([breed, kgs])

        # Save the file
        workbook.save(FILE_PATH)

        return jsonify({"message": "Entry successfully saved"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@waste_cocoons_bp.route('/waste_cocoons/exit', methods=['POST'])
def waste_cocoons_exit():
    data = request.get_json()
    breed = data.get('breed')
    kgs_to_remove = float(data.get('kgs'))

    if not breed or not kgs_to_remove:
        return jsonify({"error": "Breed and kgs are required"}), 400

    try:
        workbook = openpyxl.load_workbook(FILE_PATH)
        sheet = workbook.active

        # Track if breed was found and updated
        breed_found = False
        total_kgs_removed = 0

        for row in sheet.iter_rows(min_row=2):
            if row[0].value == breed:
                current_kgs = float(row[1].value)
                if kgs_to_remove > current_kgs:
                    return jsonify({"error": f"Cannot remove {kgs_to_remove} kgs. Only {current_kgs} kgs available for {breed}"}), 400
                # Update the kgs in the Excel file
                row[1].value = current_kgs - kgs_to_remove
                total_kgs_removed = kgs_to_remove
                breed_found = True
                break

        if not breed_found:
            return jsonify({"error": f"Breed {breed} not found in the records"}), 404

        # Save the changes to the Excel file
        workbook.save(FILE_PATH)

        return jsonify({"message": f"Successfully removed {total_kgs_removed} kgs from {breed}"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@waste_cocoons_bp.route('/waste_cocoons/available', methods=['GET'])
def get_available_waste_cocoons():
    try:
        workbook = openpyxl.load_workbook(FILE_PATH)
        sheet = workbook.active

        # Dictionary to hold available kgs by breed
        breed_totals = {}

        # Calculate the total available kgs of waste cocoons by breed
        for row in sheet.iter_rows(min_row=2, values_only=True):
            breed = row[0]  # Assuming breed is in the first column
            kgs = float(row[1])  # Assuming kgs is in the second column

            if breed in breed_totals:
                breed_totals[breed] += kgs
            else:
                breed_totals[breed] = kgs

        return jsonify(breed_totals), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
