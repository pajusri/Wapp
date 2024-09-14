from flask import Flask, Blueprint, request, jsonify, send_file
import openpyxl
from jinja2 import Environment, FileSystemLoader
from weasyprint import HTML
import os
from num2words import num2words

# Initialize Flask app
app = Flask(__name__)

# Paths
FILE_PATH = '/Users/admin/APP1/customer_requests.xlsx'
PDF_PATH = '/Users/admin/APP1/generated_pdfs/'
TEMPLATE_PATH = '/Users/admin/APP1/templates/'

# Ensure the PDF directory exists
if not os.path.exists(PDF_PATH):
    os.makedirs(PDF_PATH)

# Ensure the template directory exists
if not os.path.exists(TEMPLATE_PATH):
    os.makedirs(TEMPLATE_PATH)

# Initialize Jinja2 environment
env = Environment(loader=FileSystemLoader(TEMPLATE_PATH))

# Define the Blueprint for billing
billing_bp = Blueprint('billing', __name__)

# Function to load or create the Excel workbook
def load_or_create_workbook():
    try:
        if os.path.exists(FILE_PATH):
            workbook = openpyxl.load_workbook(FILE_PATH)
        else:
            workbook = openpyxl.Workbook()
            # Create Requests sheet
            sheet = workbook.active
            sheet.title = "Requests"
            sheet.append(['Chakki Center', 'Eggs Requested', 'Delivery Date', 'Hatching Date', 'Delivered'])
            # Create Billing sheet
            billing_sheet = workbook.create_sheet("Billing")
            billing_sheet.append(['Chakki Center', 'Eggs Requested', 'Delivery Date', 'Hatching Date', 'Status'])
            workbook.save(FILE_PATH)
        return workbook
    except Exception as e:
        print(f"Error loading or creating workbook: {str(e)}")
        raise

# Route to get billing entries
@billing_bp.route('/billing-entries', methods=['GET'])
def get_billing_entries():
    try:
        # Load the workbook and the Billing sheet
        workbook = load_or_create_workbook()
        billing_sheet = workbook['Billing']

        # Collect all rows from the Billing sheet
        billing_entries = []
        for row in billing_sheet.iter_rows(min_row=2, values_only=True):
            billing_entries.append({
                'chakkiCenter': row[0],
                'eggsRequested': row[1],
                'deliveryDate': row[2],
                'hatchingDate': row[3],
                'status': row[4]
            })

        return jsonify(billing_entries), 200
    except Exception as e:
        print(f"Error fetching billing entries: {str(e)}")
        return jsonify({'error': str(e)}), 500

# Route to delete a billing entry
@billing_bp.route('/billing/delete', methods=['POST'])
def delete_billing_entry():
    try:
        data = request.get_json()
        chakki_center = data.get('chakkiCenter')

        workbook = load_or_create_workbook()
        billing_sheet = workbook['Billing']

        row_to_delete = None
        for row in billing_sheet.iter_rows(min_row=2):
            if row[0].value == chakki_center:
                row_to_delete = row
                break

        if row_to_delete:
            billing_sheet.delete_rows(row_to_delete[0].row)
            workbook.save(FILE_PATH)
            return jsonify({'message': f'Billing entry for {chakki_center} deleted successfully'}), 200
        else:
            return jsonify({'error': f'Billing entry for {chakki_center} not found'}), 404
    except Exception as e:
        print(f"Error deleting billing entry: {str(e)}")
        return jsonify({'error': str(e)}), 500

# Route to generate a PDF bill
@billing_bp.route('/billing/generate-bill', methods=['POST'])
def generate_bill():
    try:
        data = request.get_json()
        chakki_center = data.get('chakkiCenter')
        variety = data.get('variety')
        race = data.get('race')
        lot_no = data.get('lotNo')
        laid_on = data.get('laidDate')
        ron_date = data.get('ronDate')
        eh_date = data.get('ehDate')
        quantity = int(data.get('quantity'))
        rate = float(data.get('rate'))

        # Calculate total amount
        total_amount = quantity * rate
        total_amount_formatted = "{:,.2f}".format(total_amount)
        total_amount_words = num2words(total_amount, to='cardinal', lang='en')  # Convert the total amount to words

        # Load and render the HTML template with the provided data
        template = env.get_template('bill_template.html')
        html_content = template.render(
            chakki_center=chakki_center,
            variety=variety,
            race=race,
            laid_on=laid_on,
            ron_date=ron_date,
            eh_date=eh_date,
            lot_no=lot_no,
            quantity=quantity,
            rate=rate,
            amount=f"Rs. {total_amount_formatted}",
            total_amount=total_amount_formatted,
            amount_in_words=total_amount_words.capitalize()  # Pass the words to the template
        )

        # Generate PDF from the HTML content
        pdf_filename = f"Bill_{chakki_center}.pdf"
        pdf_file_path = os.path.join(PDF_PATH, pdf_filename)
        HTML(string=html_content).write_pdf(pdf_file_path)

        # Send the generated PDF back to the client
        return send_file(pdf_file_path, as_attachment=True)

    except Exception as e:
        print(f"Error generating bill: {str(e)}")
        return jsonify({'error': str(e)}), 500

# Register the billing blueprint
app.register_blueprint(billing_bp)

# Main app runner
if __name__ == '__main__':
    app.run(debug=True)
